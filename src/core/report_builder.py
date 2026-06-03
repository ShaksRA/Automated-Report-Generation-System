"""
Excel Report Builder
Generates a professional, multi-sheet .xlsx report using openpyxl.
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from src.config import COLORS, FONTS, REPORTS_DIR, COMPANY_NAME, REPORT_TITLE

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="212121", italic=False) -> Font:
    return Font(name=FONTS["body"], bold=bold, size=size,
                color=color, italic=italic)

def _border_thin() -> Border:
    thin = Side(style="thin", color="BDBDBD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")

def _apply_header_row(ws, row_num: int, values: list, bg: str = COLORS["primary"]):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font      = _font(bold=True, size=FONTS["size_h3"], color=COLORS["white"])
        cell.fill      = _fill(bg)
        cell.alignment = _center()
        cell.border    = _border_thin()

def _apply_data_row(ws, row_num: int, values: list, alt: bool = False):
    bg = COLORS["alt_row"] if alt else COLORS["white"]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font      = _font(size=FONTS["size_body"])
        cell.fill      = _fill(bg)
        cell.alignment = _left()
        cell.border    = _border_thin()

def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _section_title(ws, row: int, col: int, text: str, span: int = 1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = _font(bold=True, size=FONTS["size_h2"], color=COLORS["primary"])
    cell.alignment = _left()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────────

def _build_cover_sheet(wb: Workbook, data: dict):
    ws = wb.active
    ws.title = "📊 Cover"
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height  = 10
    ws.row_dimensions[2].height  = 60
    ws.row_dimensions[3].height  = 30
    ws.row_dimensions[4].height  = 20
    ws.row_dimensions[5].height  = 20
    ws.row_dimensions[6].height  = 60
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 25

    # Title banner
    ws.merge_cells("B2:C2")
    title_cell = ws["B2"]
    title_cell.value     = COMPANY_NAME
    title_cell.font      = Font(name=FONTS["header"], bold=True,
                                size=28, color=COLORS["white"])
    title_cell.fill      = _fill(COLORS["primary"])
    title_cell.alignment = _center()

    ws.merge_cells("B3:C3")
    sub_cell = ws["B3"]
    sub_cell.value     = REPORT_TITLE
    sub_cell.font      = Font(name=FONTS["header"], bold=True,
                              size=16, color=COLORS["primary"])
    sub_cell.alignment = _center()

    ws.merge_cells("B4:C4")
    date_cell = ws["B4"]
    date_cell.value     = f"Generated: {datetime.now().strftime('%B %d, %Y  %H:%M')}"
    date_cell.font      = _font(italic=True, color="757575", size=11)
    date_cell.alignment = _center()

    # KPI summary cards on cover
    kpis = data["kpis"]
    ws["B6"] = "📌 KEY PERFORMANCE INDICATORS"
    ws["B6"].font      = _font(bold=True, size=FONTS["size_h2"], color=COLORS["primary"])
    ws["B6"].alignment = _left()

    headers = ["Metric", "Value"]
    _apply_header_row(ws, 7, headers, COLORS["secondary"])
    for i, (_, row) in enumerate(kpis.iterrows()):
        _apply_data_row(ws, 8 + i, [row["metric"], row["value"]], alt=i % 2 == 1)

    ws.row_dimensions[6].height = 25
    _set_col_widths(ws, [3, 40, 25])


def _build_executive_summary(wb: Workbook, data: dict):
    ws = wb.create_sheet("📋 Executive Summary")
    ws.sheet_view.showGridLines = False

    headers_info = [
        ("REPORT OVERVIEW", "This automated weekly report covers sales performance, "
         "regional breakdown, product analysis, and inventory status. "
         "All figures reflect completed transactions only."),
        ("DATA SOURCES",    "Sales exports (sales_data.csv), Customer master (customers.csv), "
         "Inventory snapshot (inventory.csv). Data pipeline: Pandas ETL → KPI Engine → Excel."),
        ("REPORT PERIOD",   datetime.now().strftime("Week ending %B %d, %Y")),
        ("GENERATED BY",    "Automated Report Generation System v1.0  |  Python · Pandas · openpyxl · REST API"),
    ]

    ws.merge_cells("A1:D1")
    ws["A1"].value     = "EXECUTIVE SUMMARY"
    ws["A1"].font      = Font(name=FONTS["header"], bold=True, size=18, color=COLORS["white"])
    ws["A1"].fill      = _fill(COLORS["primary"])
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 40

    for i, (label, text) in enumerate(headers_info):
        r = i * 2 + 3
        ws.merge_cells(f"A{r}:D{r}")
        ws[f"A{r}"].value     = label
        ws[f"A{r}"].font      = _font(bold=True, size=11, color=COLORS["white"])
        ws[f"A{r}"].fill      = _fill(COLORS["secondary"])
        ws[f"A{r}"].alignment = _left()
        ws.row_dimensions[r].height = 20

        ws.merge_cells(f"A{r+1}:D{r+1}")
        ws[f"A{r+1}"].value     = text
        ws[f"A{r+1}"].font      = _font(size=10)
        ws[f"A{r+1}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r + 1].height = 35

    # KPI table
    kpis = data["kpis"]
    start = 12
    _section_title(ws, start, 1, "Key Metrics at a Glance", span=4)
    _apply_header_row(ws, start + 1, ["Metric", "Value", "", ""], COLORS["accent"])
    for i, (_, row) in enumerate(kpis.iterrows()):
        _apply_data_row(ws, start + 2 + i,
                        [row["metric"], row["value"], "", ""], alt=i % 2 == 1)

    _set_col_widths(ws, [28, 20, 20, 20])


def _build_sales_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("💰 Sales Data")
    ws.freeze_panes = "A2"

    df = data["sales"][["order_id", "date", "region", "salesperson",
                         "product", "category", "quantity", "unit_price",
                         "discount", "net_revenue", "status", "segment"]]

    headers = ["Order ID", "Date", "Region", "Salesperson", "Product",
               "Category", "Qty", "Unit Price", "Discount", "Net Revenue",
               "Status", "Segment"]
    _apply_header_row(ws, 1, headers)

    for i, row in enumerate(df.itertuples(index=False), 2):
        vals = list(row)
        vals[1] = str(vals[1])[:10]   # date → string
        vals[7] = round(float(vals[7]), 2)
        vals[9] = round(float(vals[9]), 2)
        alt = i % 2 == 0
        bg  = COLORS["alt_row"] if alt else COLORS["white"]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font      = _font(size=FONTS["size_body"])
            cell.fill      = _fill(bg)
            cell.alignment = _left()
            cell.border    = _border_thin()

    widths = [14, 12, 10, 16, 12, 12, 6, 12, 10, 14, 12, 12]
    _set_col_widths(ws, widths)


def _build_regional_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("🗺 Regional Analysis")
    ws.sheet_view.showGridLines = False

    df = data["regional"]
    _section_title(ws, 1, 1, "Revenue by Region", span=6)
    headers = ["Region", "Orders", "Revenue ($)", "Units Sold",
               "Customers", "Avg Order Value ($)", "Revenue Share (%)"]
    _apply_header_row(ws, 2, headers)

    for i, (_, row) in enumerate(df.iterrows(), 3):
        vals = [
            row["region"],
            int(row["orders"]),
            round(row["revenue"], 2),
            int(row["units"]),
            int(row["customers"]),
            round(row["avg_order_value"], 2),
            row["revenue_share"],
        ]
        _apply_data_row(ws, i, vals, alt=i % 2 == 0)

    # Totals row
    total_row = len(df) + 3
    totals = [
        "TOTAL",
        f"=SUM(B3:B{total_row - 1})",
        f"=SUM(C3:C{total_row - 1})",
        f"=SUM(D3:D{total_row - 1})",
        "",
        "",
        "100%",
    ]
    for col, val in enumerate(totals, 1):
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font      = _font(bold=True, size=10, color=COLORS["white"])
        cell.fill      = _fill(COLORS["primary"])
        cell.alignment = _center()
        cell.border    = _border_thin()

    # Bar chart
    chart = BarChart()
    chart.type    = "col"
    chart.title   = "Revenue by Region"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Region"
    chart.style  = 10
    chart.width  = 18
    chart.height = 12

    data_ref  = Reference(ws, min_col=3, min_row=2, max_row=len(df) + 2)
    cats      = Reference(ws, min_col=1, min_row=3, max_row=len(df) + 2)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{total_row + 3}")

    _set_col_widths(ws, [14, 10, 15, 12, 12, 18, 16])


def _build_product_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("📦 Product Performance")
    ws.sheet_view.showGridLines = False

    df = data["product_perf"]
    _section_title(ws, 1, 1, "Product Performance Analysis", span=6)
    headers = ["Product", "Orders", "Units Sold", "Gross Revenue ($)",
               "Net Revenue ($)", "Avg Discount (%)", "Margin (%)"]
    _apply_header_row(ws, 2, headers)

    for i, (_, row) in enumerate(df.iterrows(), 3):
        vals = [
            row["product"],
            int(row["orders"]),
            int(row["units"]),
            round(row["gross_revenue"], 2),
            round(row["net_revenue"], 2),
            f"{row['avg_discount'] * 100:.1f}",
            f"{row['margin_pct']:.1f}",
        ]
        _apply_data_row(ws, i, vals, alt=i % 2 == 0)

    # Pie chart for net revenue
    chart = PieChart()
    chart.title  = "Net Revenue by Product"
    chart.style  = 10
    chart.width  = 16
    chart.height = 12

    n = len(df)
    data_ref = Reference(ws, min_col=5, min_row=2, max_row=n + 2)
    cats     = Reference(ws, min_col=1, min_row=3, max_row=n + 2)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = None
    ws.add_chart(chart, f"A{n + 6}")

    _set_col_widths(ws, [14, 10, 12, 18, 16, 16, 12])


def _build_salespeople_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("🏆 Sales Team")
    ws.sheet_view.showGridLines = False

    df = data["top_sales"]
    _section_title(ws, 1, 1, "Sales Team Leaderboard", span=7)
    headers = ["Rank", "Salesperson", "Orders", "Revenue ($)",
               "Units Sold", "Customers", "Rev / Order ($)"]
    _apply_header_row(ws, 2, headers)

    for i, (_, row) in enumerate(df.iterrows(), 3):
        # Gold / Silver / Bronze colours for top 3
        if row["rank"] == 1:
            bg = "FFD700"
        elif row["rank"] == 2:
            bg = "C0C0C0"
        elif row["rank"] == 3:
            bg = "CD7F32"
        else:
            bg = COLORS["alt_row"] if i % 2 == 0 else COLORS["white"]

        vals = [
            int(row["rank"]),
            row["salesperson"],
            int(row["orders"]),
            round(row["revenue"], 2),
            int(row["units"]),
            int(row["customers"]),
            round(row["revenue_per_order"], 2),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font      = _font(bold=(row["rank"] <= 3), size=10)
            cell.fill      = _fill(bg)
            cell.alignment = _left()
            cell.border    = _border_thin()

    _set_col_widths(ws, [8, 18, 10, 15, 12, 12, 15])


def _build_trend_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("📈 Daily Trend")
    ws.sheet_view.showGridLines = False

    df = data["daily_trend"]
    _section_title(ws, 1, 1, "Daily Sales Trend", span=5)
    headers = ["Date", "Orders", "Revenue ($)", "Units Sold", "Cumulative Revenue ($)"]
    _apply_header_row(ws, 2, headers)

    for i, (_, row) in enumerate(df.iterrows(), 3):
        vals = [
            row["date_str"],
            int(row["orders"]),
            round(row["revenue"], 2),
            int(row["units"]),
            round(row["cumulative_revenue"], 2),
        ]
        _apply_data_row(ws, i, vals, alt=i % 2 == 0)

    # Line chart
    chart = LineChart()
    chart.title        = "Daily Revenue Trend"
    chart.style        = 10
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Date"
    chart.width  = 22
    chart.height = 14

    n        = len(df)
    rev_ref  = Reference(ws, min_col=3, min_row=2, max_row=n + 2)
    cum_ref  = Reference(ws, min_col=5, min_row=2, max_row=n + 2)
    cats     = Reference(ws, min_col=1, min_row=3, max_row=n + 2)
    chart.add_data(rev_ref, titles_from_data=True)
    chart.add_data(cum_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{n + 6}")

    _set_col_widths(ws, [14, 10, 15, 12, 22])


def _build_inventory_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("🏭 Inventory")
    ws.sheet_view.showGridLines = False

    df = data["inventory_kpi"]
    _section_title(ws, 1, 1, "Inventory Status & Analysis", span=8)
    headers = ["Product", "Category", "SKU", "Current Stock",
               "Reorder Level", "Status", "Stock Value ($)", "Potential Rev ($)"]
    _apply_header_row(ws, 2, headers)

    for i, (_, row) in enumerate(df.iterrows(), 3):
        low = bool(row["reorder_needed"])
        bg  = "FFEBEE" if low else (COLORS["alt_row"] if i % 2 == 0 else COLORS["white"])
        vals = [
            row["product_name"],
            row["category"],
            row["sku"],
            int(row["current_stock"]),
            int(row["reorder_level"]),
            row["stock_status"],
            round(row["stock_value"], 2),
            round(row["potential_revenue"], 2),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font      = _font(size=10, bold=(col == 6 and low))
            cell.fill      = _fill(bg)
            cell.alignment = _left()
            cell.border    = _border_thin()

    _set_col_widths(ws, [14, 12, 12, 14, 14, 14, 16, 18])


def _build_category_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("🏷 Category Breakdown")
    ws.sheet_view.showGridLines = False

    df = data["category"]
    _section_title(ws, 1, 1, "Sales by Category", span=4)
    headers = ["Category", "Orders", "Revenue ($)", "Units Sold", "Revenue Share (%)"]
    _apply_header_row(ws, 2, headers)

    for i, (_, row) in enumerate(df.iterrows(), 3):
        vals = [
            row["category"],
            int(row["orders"]),
            round(row["revenue"], 2),
            int(row["units"]),
            row["revenue_share"],
        ]
        _apply_data_row(ws, i, vals, alt=i % 2 == 0)

    _set_col_widths(ws, [16, 10, 15, 12, 18])


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────

def build_report(data: Dict) -> Path:
    """Build and save the Excel report. Returns the output Path."""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = REPORTS_DIR / f"weekly_report_{timestamp}.xlsx"

    logger.info("Building Excel report → %s", output_path)

    wb = Workbook()
    _build_cover_sheet(wb, data)
    _build_executive_summary(wb, data)
    _build_sales_sheet(wb, data)
    _build_regional_sheet(wb, data)
    _build_product_sheet(wb, data)
    _build_salespeople_sheet(wb, data)
    _build_trend_sheet(wb, data)
    _build_inventory_sheet(wb, data)
    _build_category_sheet(wb, data)

    wb.save(output_path)
    logger.info("Report saved: %s", output_path)
    return output_path
